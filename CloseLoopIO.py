import json  # JSON read/write for metadata and scalar files
import os  # Filesystem paths and directory creation
import glob  # Pattern-based file enumeration
import zipfile  # Create/read ZIP archives
import tempfile  # Temporary directories for extraction
import shutil  # File operations (copy/remove, make_archive)
import logging  # Logging info/warnings/errors
import subprocess  # Launch external editor for README notes
from datetime import datetime  # Timestamps for filenames and logs
from zoneinfo import ZoneInfo  # Time zone handling (Europe/Paris)

import numpy as np  # Numerical arrays and .npy IO
from matplotlib.figure import Figure  # Type hint for matplotlib Figure
from openpyxl import load_workbook, Workbook  # XLSX workbook read/write
from openpyxl.styles import Alignment, Font  # Cell styling for headers

from OZIRIIS.CloseLoop import CloseLoop  # Target class we save/reload around

logging.basicConfig(level=logging.INFO)  # Configure root logger (INFO)
logger = logging.getLogger(__name__)  # Module-level logger


class OziriisIO:
    """Save/load utility + XLSX logging for CloseLoop sessions."""  # Class purpose docstring

    def __init__(self, common_path: str = '/media/manip/4E3B-E8FD/nextcloud/These/Experimentation_bench/OZIRIIS/',savedir: str = "data", logdir: str = "log"):
        self.savedir = common_path+savedir  # Base directory where sessions are saved
        os.makedirs(common_path+savedir, exist_ok=True)  # Ensure save directory exists
        os.makedirs(common_path+logdir, exist_ok=True)  # Ensure log directory exists
        self.log_xlsx_path = os.path.join(common_path+logdir, "log_OZIRIIS.xlsx")  # Path to global XLSX journal

    # --------------------------
    # Small helpers
    # --------------------------
    def _timestamp(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")  # Timestamp (to second) for filenames

    def _build_path(self, prefix: str, extension: str = ".npy", root: str | None = None) -> str:
        root = self.savedir if root is None else root  # Default root to savedir
        return os.path.join(root, f"{prefix}_{self._timestamp()}{extension}")  # Compose path with timestamp

    def _to_jsonable(self, val):
        return val.item() if isinstance(val, np.generic) else val  # Convert numpy scalars to Python scalars
    def _get(self, obj, *names, default=None):
        """Return first existing attribute among names, else default."""
        for n in names:
            try:
                return getattr(obj, n)
            except Exception:
                pass
        return default

    def _get_cam(self, cl, name):
        """Read camera attribute `name` with fallbacks:
        cl.cam_<name> -> cl._cam_<name> -> cl.cam.<name> -> None."""
        return self._to_jsonable(
            self._get(cl, f"cam_{name}", f"_cam_{name}",
                    default=self._get(cl.cam, name))
        )
    # --------------------------
    # Saving block
    # --------------------------
    def save_metadata(self, cl, prefix: str = "metadata", root: str | None = None, extra_scalars: dict | None = None) -> str:
        metadata = {}

        # --- CloseLoop core (robust getters handle public/underscored names) ---
        metadata["gain"]               = self._to_jsonable(self._get(cl, "gain", "_gain"))
        metadata["iteration"]          = self._to_jsonable(self._get(cl, "iteration", "_iteration"))
        metadata["inversion_trunc"]    = self._to_jsonable(self._get(cl, "inversion_trunc", "_inversion_trunc"))
        metadata["IM_modal"]           = self._to_jsonable(self._get(cl, "IM_modal", "_IM_modal"))
        metadata["controlled_modes"]   = self._to_jsonable(self._get(cl, "controlled_modes", "_controlled_modes"))
        metadata["ZWFS_tag"]           = self._to_jsonable(self._get(cl, "ZWFS_tag", "_ZWFS_tag"))
        metadata["ZWFS_shift"]         = self._to_jsonable(self._get(cl, "ZWFS_shift", "_ZWFS_shift"))
        metadata["ratio_mask_psf"]     = self._to_jsonable(self._get(cl, "ratio_mask_psf", "_ratio_mask_psf"))
        metadata["precropping_data"]   = self._to_jsonable(self._get(cl, "precropping_data"))
        metadata["validpixels_crop"]   = self._to_jsonable(self._get(cl, "validpixels_crop", "_valid_pixel_crop"))
        metadata["zpf"]                = self._to_jsonable(self._get(cl, "zpf", "_zpf"))
        metadata["reconstructor_type"] = self._to_jsonable(self._get(cl, "reconstructor_type", "_reconstructor_type"))
        metadata["modal_command"]      = self._to_jsonable(self._get(cl, "modal_command", "_modal_command"))

        # optional / may be absent (keep warnings quiet by using defaults)
        optional_keys = [
            "IM_nframes", "IM_stroke", "IM_nmeasurement",
            "close_loop_nFrames", "leakage",
        ]
        for key in optional_keys:
            val = self._get(cl, key, f"_{key}", default=None)
            if val is not None:
                metadata[key] = self._to_jsonable(val)

        # --- Camera: prefer cl.cam_* mirrors, then cl.cam.<attr> ---
        for name in ["roi", "binning", "gain", "tint", "fps", "temp"]:
            try:
                metadata[f"cam_{name}"] = self._get_cam(cl, name)
            except Exception as e:
                logger.warning(f"Could not retrieve camera attribute {name}: {e}")

        # --- merge user-provided extra scalar metadata (new feature in your class) ---
        if extra_scalars:
            for k, v in extra_scalars.items():
                if k in metadata and metadata[k] != v:
                    logger.info(f"Overriding metadata key '{k}' with provided extra value")
                metadata[k] = self._to_jsonable(v)

        metadata["save_datetime"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        path = self._build_path(prefix, ".json", root)
        with open(path, "w") as f:
            json.dump(metadata, f, indent=4)
        return path

    def save_array(self, arr: np.ndarray, prefix: str, root: str | None = None) -> str:
        path = self._build_path(prefix, ".npy", root)  # Target .npy path
        np.save(path, arr)  # Save NumPy array
        return path  # Return saved path

    def save_scalar(self, val, prefix: str, root: str | None = None) -> str:
        path = self._build_path(prefix, ".json", root)  # Target JSON path for a scalar
        with open(path, "w") as f:
            json.dump({"value": self._to_jsonable(val)}, f, indent=4)  # Wrap scalar in a dict
        return path  # Return path

    def save_plot(self, fig: Figure, prefix: str, root: str | None = None) -> tuple[str, str]:
        root = self.savedir if root is None else root  # Default to savedir
        ts = self._timestamp()  # Unique suffix for figure files
        png_path = os.path.join(root, f"{prefix}_{ts}.png")  # PNG export path
        fig_path = os.path.join(root, f"{prefix}_{ts}.fig.pkl")  # Pickled Figure path
        fig.savefig(png_path, dpi=300)  # Save static PNG for quick viewing
        try:
            import pickle  # Local import to avoid hard dependency at module import
            with open(fig_path, "wb") as f:
                pickle.dump(fig, f)  # Serialize the live Figure for later editing
        except Exception as e:
            logger.warning(f"Could not pickle matplotlib figure: {e}")  # Non-fatal if pickling fails
        return png_path, fig_path  # Return both paths

    def collect_user_notes(self, root: str) -> str | None:
        """Open $EDITOR on README.txt and return its path if created."""  # Explain interactive step
        temp_txt_path = os.path.join(root, "README.txt")  # Path of the notes file
        try:
            editor = os.environ.get("EDITOR", "nano")  # Fallback to nano if EDITOR unset
            subprocess.call([editor, temp_txt_path])  # Block until editor closes
        except Exception as e:
            logger.warning(f"Could not open editor for README: {e}")  # Non-fatal if no editor
        return temp_txt_path if os.path.exists(temp_txt_path) else None  # Return only if file exists

    def save_results(self, results: dict, session_dir: str) -> None:
        results_dir = os.path.join(session_dir, "results")  # Subfolder for results
        os.makedirs(results_dir, exist_ok=True)  # Ensure it exists
        scalar_store: dict[str, object] = {}  # Aggregates scalar results → one JSON file
        for key, val in (results or {}).items():  # Iterate provided results
            try:
                if isinstance(val, np.ndarray):  # Arrays -> .npy
                    self.save_array(val, key, results_dir)
                elif isinstance(val, (int, float, str, bool, np.generic)):  # Scalars -> collect
                    scalar_store[key] = self._to_jsonable(val)
                elif isinstance(val, Figure):  # Matplotlib figure -> save PNG + pickle
                    self.save_plot(val, key, results_dir)
                else:
                    logger.warning(f"Unsupported result type for {key}: {type(val)}")  # Ignore unknown types
            except Exception as e:
                logger.warning(f"Failed to save result {key}: {e}")  # Continue saving others
        if scalar_store:  # Write scalars.json only if non-empty
            with open(os.path.join(results_dir, "scalars.json"), "w") as f:
                json.dump(scalar_store, f, indent=4)

    def _extract_scalar_results_from_input(self, results: dict | None) -> dict[str, object]:
        """Return only scalar-like values from `results` dict (no disk IO).
        Keys are prefixed with 'result.' here, so downstream must NOT re-prefix.
        """
        out: dict[str, object] = {}  # Output map: 'result.key' -> scalar
        if not results:
            return out  # Nothing to extract

        def _as_scalar(v):  # Helper: extract scalar from value/list/array
            if isinstance(v, (int, float, str, bool, np.generic)):
                return self._to_jsonable(v)  # Direct scalar
            if isinstance(v, (list, tuple, np.ndarray)) and len(v) == 1:
                vv = v[0]  # Single-element list/array → scalar
                if isinstance(vv, (int, float, str, bool, np.generic)):
                    return self._to_jsonable(vv)
            return None  # Not scalar-like

        for k, v in results.items():  # Prefix keys and collect
            s = _as_scalar(v)
            if s is not None:
                out[f"result.{k}"] = s
        return out

    # --------------------------
    # XLSX log helpers
    # --------------------------
    def _flatten_scalars(self, d: dict, parent: str = "") -> dict:
        """Flatten nested dict into dotted keys and keep only scalar-like values."""
        flat = {}  # Output flattened dict
        for k, v in d.items():  # Walk dict recursively
            key = f"{parent}.{k}" if parent else k  # Compose dotted key
            if isinstance(v, dict):
                flat.update(self._flatten_scalars(v, key))  # Recurse
            elif isinstance(v, (int, float, str, bool, np.generic)):
                flat[key] = self._to_jsonable(v)  # Keep basic scalars
            elif isinstance(v, tuple):
                flat[key] = str(v)  # Represent tuples as strings for logging
            elif isinstance(v, (list, np.ndarray)):
                try:
                    if len(v) == 1:  # Single-element sequences → scalar
                        vv = v[0]
                        if isinstance(vv, (int, float, str, bool, np.generic)):
                            flat[key] = self._to_jsonable(vv)
                except Exception:
                    pass  # Ignore non-sized objects
        return flat  # Return flattened map

    def _family_for_key(self, key: str) -> str:
        """Map a column key to a human-friendly group name for row-1 headers."""
        if key in ("exp_index", "log_modification_date", "zip_path"):
            return "Base"  # Session basics
        if key in ("details", "comment"):
            return "Notes"  # Free-form notes

        # Group definitions for metadata
        close_loop = {"gain", "iteration", "controlled_modes", "close_loop_nFrames", "leakage"}
        im_params = {"IM_nframes", "IM_stroke", "IM_nmeasurement", "IM_modal"}
        recon = {"reconstructor_type", "modal_command", "inversion_trunc"}
        zernike = {"ZWFS_tag", "ZWFS_shift", "ratio_mask_psf", "zpf"}

        if key in close_loop:
            return "Loop parameters"
        if key in im_params:
            return "IM parameters"
        if key in recon:
            return "Reconstruction"
        if key in zernike:
            return "ZWFS parameters"
        if key.startswith("cam_"):
            return "Camera"  # Camera.* keys
        if key.startswith("dm_"):
            return "DM"  # DM.* keys (if used)
        if key.startswith("result.") or key.startswith("result_"):
            return "Results"  # Scalar results
        if key.startswith("extra.") or key.startswith("extra_"):
            return "Extra"  # Extra context fields
        return "Metadata other"  # Everything else

    def _row_has_content(self, ws, r: int) -> bool:
        return any((c.value not in (None, "")) for c in ws[r])  # True if any cell in row is non-empty

    def _next_data_row(self, ws, header_rows: int = 2) -> int:
        last = 0  # Track last non-empty row
        for r in range(ws.max_row, 0, -1):  # Scan upwards from bottom
            if self._row_has_content(ws, r):
                last = r
                break
        return max(header_rows + 1, last + 1)  # First data row is after header

    # --- Helpers to incorporate user-provided extra metadata ---
    def _safe_key(self, key: str) -> str:
        """Return a filesystem-safe version of a key (used in filenames)."""
        return "".join(ch if (ch.isalnum() or ch in ("_", "-")) else "_" for ch in key)

    def _split_meta_extra(self, meta_extra: dict | None) -> tuple[dict, dict]:
        """Split extra metadata dict into (scalars, arrays).
        - Scalars: int/float/str/bool/np.generic or single-element lists/arrays.
        - Arrays: numpy arrays OR lists/tuples of length != 1 (converted to np.array).
        Nested dicts are flattened using dotted keys.
        """
        scalars: dict[str, object] = {}
        arrays: dict[str, np.ndarray] = {}
        if not meta_extra:
            return scalars, arrays

        def walk(prefix: str, val):
            # Determine full key name
            k = prefix
            # Classify value
            if isinstance(val, (int, float, str, bool, np.generic)):
                scalars[k] = self._to_jsonable(val)
            elif isinstance(val, dict):
                for subk, subv in val.items():
                    subname = f"{k}.{subk}" if k else str(subk)
                    walk(subname, subv)
            elif isinstance(val, (list, tuple, np.ndarray)):
                try:
                    # Treat single element list/array as scalar if scalar-like
                    if len(val) == 1:
                        vv = val[0]
                        if isinstance(vv, (int, float, str, bool, np.generic)):
                            scalars[k] = self._to_jsonable(vv)
                            return
                except Exception:
                    pass
                # Otherwise, store as array
                arrays[k] = np.asarray(val)
            else:
                logger.warning(f"Unsupported extra metadata type for key '{k}': {type(val)}")

        for key, value in meta_extra.items():
            walk(str(key), value)
        return scalars, arrays

    # --- Helpers to manage per-day sheets and manual additions ---
    def _get_or_create_sheet_named(self, wb: Workbook, name: str):
        """Return a worksheet named `name`, creating it if needed.
        If workbook is pristine (single default 'Sheet'), rename it.
        """
        if name in wb.sheetnames:
            return wb[name]  # Reuse existing sheet
        if (
            len(wb.sheetnames) == 1
            and wb.active.title == "Sheet"
            and wb.active.max_row == 1
            and wb.active.max_column == 1
        ):
            ws = wb.active  # Use default sheet
            ws.title = name  # Rename to requested name
            return ws
        return wb.create_sheet(title=name)  # Otherwise create a new sheet

    def _date_sheet_indices(self, wb: Workbook, base: str) -> list[int]:
        """Return sorted list of indices for sheets of a given day.
        0 represents the base sheet named exactly `base`.
        Positive integers represent suffixed sheets like `base_2`.
        """
        idxs: list[int] = []  # Collected indices
        for title in wb.sheetnames:  # Inspect all sheet names
            if title == base:
                idxs.append(0)  # Base exists
            elif title.startswith(base + "_"):
                suf = title[len(base) + 1 :]  # Extract numeric suffix
                if suf.isdigit():
                    idxs.append(int(suf))  # Add numeric suffix
        return sorted(idxs)  # Return ascending order

    def _select_or_create_sheet(self, wb: Workbook, base: str, manual_new: bool):
        """Pick the worksheet for date `base`.
        - Default: use latest existing sheet; if none, create `base`.
        - manual_new=True: create a *new* sheet for that day.
          If a plain `base` sheet exists, rename it to `base_1` and create `base_2`.
          Otherwise create `base_{N+1}` with N the current max index.
        Returns (worksheet, resolved_title).
        """
        idxs = self._date_sheet_indices(wb, base)  # Existing indices for that date
        if not manual_new:
            if not idxs:  # No sheet for that date yet
                ws = self._get_or_create_sheet_named(wb, base)
                return ws, ws.title
            chosen = max(idxs)  # Use latest sheet index
            name = base if chosen == 0 else f"{base}_{chosen}"  # Resolve sheet name
            return wb[name], name
        # manual_new requested
        if not idxs:  # No existing -> create base
            ws = self._get_or_create_sheet_named(wb, base)
            return ws, ws.title
        if 0 in idxs and f"{base}_1" not in wb.sheetnames:  # If base exists and _1 not present
            wb[base].title = f"{base}_1"  # Rename base to suffix _1
            idxs = self._date_sheet_indices(wb, base)  # Refresh indices after rename
        existing_pos = [i for i in idxs if i > 0]  # All suffixed sheets
        next_idx = (max(existing_pos) if existing_pos else 1) + 1  # Next index to allocate
        name = f"{base}_{next_idx}"  # New sheet name
        ws = wb.create_sheet(title=name)  # Create sheet
        return ws, name  # Return worksheet and title

    def log_to_xlsx(self, zip_path: str, metadata_path: str,
                    results_scalars: dict, extra: dict, comment: str,
                    new_sheet: bool = False) -> None:
        """Append one row per save in an Excel journal, grouped headers on 2 rows."""
        now = datetime.now(ZoneInfo("Europe/Paris"))  # Current time in Paris
        sheet_name = now.strftime("%Y-%m-%d")  # Sheet label for the day
        log_modification_date = now.strftime("%Y-%m-%d %H:%M:%S")  # Log entry timestamp

        if os.path.exists(self.log_xlsx_path):  # Open existing workbook…
            wb = load_workbook(self.log_xlsx_path)
        else:
            wb = Workbook()  # …or create new one

        ws, resolved_title = self._select_or_create_sheet(wb, sheet_name, manual_new=new_sheet)  # Resolve sheet

        with open(metadata_path, "r") as f:  # Read metadata JSON
            meta_raw = json.load(f)
        meta = self._flatten_scalars(meta_raw)  # Flatten to dotted scalar dict

        # Ordered families for metadata columns
        fam_order = [
            ("close_loop", ["gain", "iteration", "controlled_modes", "close_loop_nFrames", "leakage"]),
            ("interaction_matrix", ["IM_nframes", "IM_stroke", "IM_nmeasurement", "IM_modal"]),
            ("reconstruction", ["reconstructor_type", "modal_command", "inversion_trunc"]),
            ("zernike", ["ZWFS_tag", "ZWFS_shift", "ratio_mask_psf", "zpf"]),
            ("camera", None),  # Any cam_* keys (alphabetical)
            ("dm", None),      # Any dm_* keys (alphabetical)
            ("metadata_other", None),  # Remaining keys
        ]
        meta_keys = set(meta.keys())  # Present keys
        meta_cols: list[str] = []  # Final ordered metadata columns
        explicit_keys = set()  # Track keys already placed by explicit lists
        group_for_col: dict[str, str] = {}  # Column -> family group

        for fam, keys in fam_order:  # Build metadata column order
            if keys is not None:  # Explicit key list for this family
                for k in keys:
                    if k in meta:
                        meta_cols.append(k)  # Append key if present
                        group_for_col[k] = fam  # Group mapping
                        explicit_keys.add(k)  # Mark as used
            else:  # Prefix-based or leftover families
                if fam == "camera":
                    ks = sorted([k for k in meta_keys if k.startswith("cam_")])  # All cam_*
                elif fam == "dm":
                    ks = sorted([k for k in meta_keys if k.startswith("dm_")])  # All dm_*
                else:  # metadata_other
                    ks = sorted([k for k in meta_keys if k not in explicit_keys and not k.startswith("cam_") and not k.startswith("dm_")])
                for k in ks:
                    meta_cols.append(k)  # Append discovered key
                    group_for_col[k] = fam  # Map to its family
                    explicit_keys.add(k)  # Mark as used

        results_scalars = results_scalars or {}  # Ensure dict
        res_cols = sorted(results_scalars.keys())  # Already prefixed as 'result.'
        for k in res_cols:
            group_for_col[k] = "results"  # Group label for results

        extra = extra or {}  # Ensure dict
        extra_cols = [f"extra.{k}" for k in sorted(extra.keys())]  # Prefix extra keys
        for k in extra_cols:
            group_for_col[k] = "extra"  # Group label for extra

        header_cols = [  # Final header order (2 fixed + path, meta, results, extra, two notes)
            "exp_index", "log_modification_date", "zip_path"
        ] + meta_cols + res_cols + extra_cols + ["details", "comment"]

        existing = []  # Existing header (row 2) if any
        if ws.max_row >= 2:
            existing = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]  # Read header row
            existing = [x for x in existing if x not in (None, "")]  # Drop empties
        has_header = (len(existing) > 0 and existing[0] == "exp_index")  # Detect valid header

        def write_header(cols: list[str]):  # (Re)write the 2 header rows
            for rng in list(ws.merged_cells.ranges):  # Clear any existing merges in row 1
                ws.unmerge_cells(str(rng))
            ws.row_dimensions[1].height = 24  # Row 1 height
            ws.row_dimensions[2].height = 20  # Row 2 height
            for j, name in enumerate(cols, start=1):  # Row 2: column names
                c = ws.cell(row=2, column=j)
                c.value = name
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cur_group = None  # Track current group block
            start = 1  # Start column of current block
            for j, name in enumerate(cols, start=1):  # Row 1: merged group titles
                g = self._family_for_key(name)  # Determine family for this column
                if cur_group is None:  # First column
                    cur_group, start = g, j
                elif g != cur_group:  # Group change → close previous merge
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=j - 1)
                    top_left = ws.cell(row=1, column=start)
                    top_left.value = cur_group
                    top_left.font = Font(bold=True)
                    top_left.alignment = Alignment(horizontal="center", vertical="center")
                    cur_group, start = g, j  # Start new group
            if cur_group:  # Close last group to the end
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=len(cols))
                top_left = ws.cell(row=1, column=start)
                top_left.value = cur_group
                top_left.font = Font(bold=True)
                top_left.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A3"  # Freeze panes below headers

        if not has_header:  # New sheet → write header from scratch
            write_header(header_cols)
        else:
            # Re-read the current header row (row 2)
            existing = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
            existing = [x for x in existing if x not in (None, "")]

            # Where to insert new columns? -> just before the first of 'details' or 'comment'
            def _find_anchor(cols):
                for name in ("details", "comment"):
                    if name in cols:
                        # openpyxl is 1-based for column indices
                        return cols.index(name) + 1
                # If Notes are not present yet, append at the end
                return ws.max_column + 1

            anchor = _find_anchor(existing)

            # Compute which columns are missing (ignore Notes; we’ll ensure them later)
            missing = [c for c in header_cols if c not in existing and c not in ("details", "comment")]

            if missing:
                # Insert a span just before Notes and write the new header names there
                ws.insert_cols(anchor, amount=len(missing))
                for j, name in enumerate(missing, start=anchor):
                    ws.cell(row=2, column=j).value = name

            # Make sure 'details' and 'comment' exist (at the end) in that order
            def _header_values():
                return [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

            hdr = _header_values()
            for name in ("details", "comment"):
                if name not in hdr:
                    ws.cell(row=2, column=ws.max_column + 1).value = name
                    hdr = _header_values()  # refresh after each append

            # Rebuild the merged group titles (row 1) over the final physical order
            final_cols = _header_values()
            write_header(final_cols)
            header_cols = final_cols


        details_text = ""  # Default README content
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:  # Open session zip
                names = zf.namelist()  # Content list
                readme_names = [n for n in names if n.endswith("README.txt")]  # Find README path
                if readme_names:
                    raw = zf.read(readme_names[0])  # Read bytes
                    try:
                        details_text = raw.decode("utf-8")  # UTF-8 first
                    except UnicodeDecodeError:
                        details_text = raw.decode("latin-1")  # Fallback encoding
        except Exception as e:
            logger.warning(f"Impossible to read README.txt from the zip: {e}")  # Non-fatal

        values = {  # Values for this row (matched by header name)
            "exp_index": None,  # Filled after we find target row index
            "log_modification_date": log_modification_date,  # When the log entry was created
            "zip_path": os.path.abspath(zip_path),  # Absolute path to the saved zip
            "details": details_text,  # README content
            "comment": comment or "",  # User-provided comment
        }
        for k in meta_cols:  # Inject metadata values
            if k in meta:
                values[k] = meta[k]
        for col, val in results_scalars.items():  # Inject results scalars
            values[col] = val
        for k, v in extra.items():  # Inject extra fields (prefixed extra.)
            values[f"extra.{k}"] = v

        row = self._next_data_row(ws, header_rows=2)  # First free row after header
        exp_index = max(1, row - 2)  # Daily experiment index (1-based)
        values["exp_index"] = exp_index  # Set exp_index

        for j, colname in enumerate(header_cols, start=1):  # Write all cells in order
            ws.cell(row=row, column=j).value = values.get(colname, "")

        wb.save(self.log_xlsx_path)  # Persist workbook to disk
        logger.info(f"Line added in {self.log_xlsx_path} → sheet {sheet_name}, line {row}")  # Log success

    # --------------------------
    # save_all (with logging)
    # --------------------------
    def save_all(self, cl, prefix: str = "run", results: dict | None = None, meta_extra: dict | None = None,
                 compress: bool = True, extra_log: dict | None = None, comment: str = "",
                 log_new_sheet: bool = False):
        timestamp = self._timestamp()  # Unique session folder name
        session_dir = os.path.join(self.savedir, timestamp)  # Path to the session folder
        os.makedirs(session_dir, exist_ok=True)  # Create session folder

        meta_dir = os.path.join(session_dir, "metadata")  # Subfolder for metadata arrays & JSON
        os.makedirs(meta_dir, exist_ok=True)  # Ensure it exists

        # Split user-provided extra metadata into scalars vs arrays
        extra_meta_scalars, extra_meta_arrays = self._split_meta_extra(meta_extra)

        paths: dict[str, str] = {}  # Track saved file paths
        # Save metadata JSON, including any extra scalar metadata provided by the user
        paths["metadata"] = self.save_metadata(cl, prefix + "_metadata", meta_dir, extra_scalars=extra_meta_scalars)

        # Core arrays (saved into metadata/ for locality)
        paths["history_cmd"]    = self.save_array(cl.history_cmd,    prefix + "_cmd",       meta_dir)
        paths["history_signal"] = self.save_array(cl.history_signal, prefix + "_signal",    meta_dir)
        paths["IM"]             = self.save_array(cl.IM,             prefix + "_IM",        meta_dir)
        paths["pupil"]          = self.save_array(cl.pupil,          prefix + "_pupil",     meta_dir)
        paths["submask"]        = self.save_array(cl.submask,        prefix + "_submask",   meta_dir)

        # Optional arrays (guard against absence)
        try:
            if getattr(cl, "reconstructed_phase", None) is not None:
                paths["reconstructed_phase"] = self.save_array(cl.reconstructed_phase, prefix + "_recon", meta_dir)
        except Exception as e:
            logger.warning(f"Could not save reconstructed_phase: {e}")
        try:
            if getattr(cl, "injected_perturbation", None) is not None:
                paths["injected_perturbation"] = self.save_array(cl.injected_perturbation, prefix + "_perturbation", meta_dir)
        except Exception as e:
            logger.warning(f"Could not save injected_perturbation: {e}")
        try:
            if hasattr(cl.cam, "dark") and cl.cam.dark is not None:
                paths["cam_dark"] = self.save_array(cl.cam.dark, prefix + "_dark", meta_dir)
        except Exception as e:
            logger.warning(f"Could not save camera dark: {e}")
        try:
            if hasattr(cl.dm, "IF") and cl.dm.IF is not None:
                paths["dm_IF"] = self.save_array(cl.dm.IF, prefix + "_IF", meta_dir)
        except Exception as e:
            logger.warning(f"Could not save DM IF: {e}")
        try:
            if hasattr(cl.dm, "M2C") and cl.dm.M2C is not None:
                paths["dm_M2C"] = self.save_array(cl.dm.M2C, prefix + "_M2C", meta_dir)
        except Exception as e:
            logger.warning(f"Could not save DM M2C: {e}")

        # Save any extra metadata arrays provided by the user into metadata/
        for k, arr in extra_meta_arrays.items():
            key_safe = self._safe_key(k)
            try:
                self.save_array(arr, f"{prefix}_{key_safe}", meta_dir)
            except Exception as e:
                logger.warning(f"Failed to save extra metadata array '{k}': {e}")

        if results:  # Save additional results (arrays/figures/scalars)
            self.save_results(results, session_dir)

        self.collect_user_notes(session_dir)  # Open README for user notes (optional)

        if compress:  # Create .zip archive of the entire session folder
            zip_path = shutil.make_archive(session_dir, "zip", session_dir)  # Make archive next to folder
            try:
                self.log_to_xlsx(  # Append a row to the XLSX journal
                    zip_path=zip_path,
                    metadata_path=paths["metadata"],
                    results_scalars=self._extract_scalar_results_from_input(results),
                    extra=extra_log or {},
                    comment=comment,
                    new_sheet=log_new_sheet,
                )
            except Exception as e:
                logger.warning("Log not updated: %s", e)  # Do not fail the save on logging errors
            shutil.rmtree(session_dir)  # Remove the uncompressed folder after zipping
            logger.info(f"Saved and compressed session to {zip_path}")  # Report success
            return zip_path  # Return path to zip
        else:
            logger.info(f"Saved session to {session_dir} (not compressed)")  # No compression path
            return session_dir  # Return raw folder path

    # --------------------------
    # Loading block
    # --------------------------
    def _load_json(self, path: str):
        with open(path, "r") as f:
            return json.load(f)  # Read JSON file and return dict

    def _load_latest_npy_in_dir(self, dir_path: str, stem: str):
        pattern = os.path.join(dir_path, f"{stem}_*.npy")  # Pattern for timestamped arrays
        candidates = sorted(glob.glob(pattern))  # List matching files
        return np.load(candidates[-1], allow_pickle=True) if candidates else None  # Load newest or None

    def load_from_meta_then_root(self, stem: str):
        arr = self._load_latest_npy_in_dir(self._meta_dir, stem)  # Try metadata/ first
        if arr is None and self._meta_dir != self._session_root:  # Fallback to session root
            arr = self._load_latest_npy_in_dir(self._session_root, stem)
        return arr  # Return array or None

    def load_session(self, zip_or_dir: str, dm, cam, keep_extracted: bool = False):
        cleanup_needed = False  # Whether to delete temp extraction dir
        if os.path.isdir(zip_or_dir):  # Input is a folder
            extracted_dir = zip_or_dir  # Use as-is
        else:  # Input is a .zip file → extract to temp
            tmpdir = tempfile.mkdtemp(prefix="oziriis_load_")  # Create temp dir
            with zipfile.ZipFile(zip_or_dir, "r") as zf:
                zf.extractall(tmpdir)  # Unpack archive
            extracted_dir = tmpdir  # Set extracted root
            cleanup_needed = not keep_extracted  # Mark for cleanup unless kept

        subdirs = [p for p in os.listdir(extracted_dir) if os.path.isdir(os.path.join(extracted_dir, p))]  # List subfolders
        session_root = os.path.join(extracted_dir, subdirs[0]) if len(subdirs) == 1 else extracted_dir  # Choose session root

        meta_dir = os.path.join(session_root, "metadata")  # Prefer metadata/ subfolder
        meta_candidates = sorted(glob.glob(os.path.join(meta_dir, "*metadata*.json"))) if os.path.isdir(meta_dir) else []  # Find metadata JSON
        if not meta_candidates:  # Backward-compat: search anywhere within session
            meta_candidates = sorted(glob.glob(os.path.join(session_root, "**", "*_metadata_*.json"), recursive=True))
            if not meta_candidates:
                meta_candidates = sorted(glob.glob(os.path.join(session_root, "**", "*metadata*.json"), recursive=True))
            if not meta_candidates:
                raise FileNotFoundError("Metadata files not found in the session")  # Nothing to load
            meta_dir = session_root  # Arrays likely live in root for old format
        self._meta_dir = meta_dir  # Remember for helper loads
        self._session_root = session_root  # Remember for helper loads
        meta_path = meta_candidates[-1]  # Use most recent metadata JSON

        base = os.path.basename(meta_path)  # Filename only
        prefix = base.split("_metadata_")[0] if "_metadata_" in base else base.replace(".json", "").replace("_metadata", "")  # Infer prefix

        metadata = self._load_json(meta_path)  # Load metadata values
        params = dict(  # Build CloseLoop constructor args (with defaults)
            gain=metadata.get("gain", 0.3),
            iteration=metadata.get("iteration", 100),
            inversion_trunc=metadata.get("inversion_trunc", 0),
            IM_modal=metadata.get("IM_modal", True),
            ZWFS_tag=metadata.get("ZWFS_tag", 1),
            ZWFS_shift=metadata.get("ZWFS_shift", 0.33),
            ratio_mask_psf=metadata.get("ratio_mask_psf", 2.2),
            precropping_data=metadata.get("precropping_data", [50, 400, 50, 400]),
            controlled_modes=metadata.get("controlled_modes", None),
            validpixels_crop=metadata.get("validpixels_crop", 1e3),
            zpf=metadata.get("zpf", 30),
            reconstructor_type=metadata.get("reconstructor_type", "asin"),
        )

        IM = self.load_from_meta_then_root(f"{prefix}_IM")  # Interaction matrix (signal-sized)
        pupil = self.load_from_meta_then_root(f"{prefix}_pupil")  # Pupil map
        submask = self.load_from_meta_then_root(f"{prefix}_submask")  # Submask map
        if IM is None or pupil is None or submask is None:
            raise FileNotFoundError("IM/pupil/submask not found : impossible to reconstruct CloseLoop")  # Critical

        cl = CloseLoop(  # Rebuild CloseLoop object
            dm=dm, cam=cam, IM=IM,
            gain=params["gain"], iteration=params["iteration"],
            inversion_trunc=params["inversion_trunc"], IM_modal=params["IM_modal"],
            ZWFS_tag=params["ZWFS_tag"], ZWFS_shift=params["ZWFS_shift"],
            ratio_mask_psf=params["ratio_mask_psf"], precropping_data=params["precropping_data"],
            controlled_modes=params["controlled_modes"], validpixels_crop=params["validpixels_crop"],
            zpf=params["zpf"], reconstructor_type=params["reconstructor_type"],
        )

        try:
            cl.pupil = pupil  # Inject saved pupil
            cl.submask = submask  # Inject saved submask
        except Exception as e:
            logger.warning(f"Impossible to inject pupil/submask: {e}")  # Not fatal but unexpected

        hist_cmd = self.load_from_meta_then_root(f"{prefix}_cmd")  # Command history
        hist_sig = self.load_from_meta_then_root(f"{prefix}_signal")  # Signal history
        recon = self.load_from_meta_then_root(f"{prefix}_recon")  # Reconstructed phase
        perturb = self.load_from_meta_then_root(f"{prefix}_perturbation")  # Injected perturbation
        if hist_cmd is not None:
            cl._history_cmd = hist_cmd  # Restore private field
        if hist_sig is not None:
            cl._history_signal = hist_sig  # Restore private field
        if recon is not None:
            cl._reconstructed_phase = recon  # Restore private field
        if perturb is not None:
            cl._injected_perturbation = perturb  # Restore private field

        cam_dark = self.load_from_meta_then_root(f"{prefix}_dark")  # Camera dark
        if cam_dark is not None:
            try:
                cl.cam.set_dark(cam_dark)  # Prefer setter if available
            except Exception:
                cl.cam.dark = cam_dark  # Fallback to direct attribute

        dm_if = self.load_from_meta_then_root(f"{prefix}_IF")  # DM influence functions
        if dm_if is not None:
            try:
                dm.IF = dm_if  # Restore IF into DM
            except Exception as e:
                logger.warning(f"Impossible to apply dm.IF: {e}")  # Non-fatal
        dm_m2c = self.load_from_meta_then_root(f"{prefix}_M2C")  # DM M2C matrix
        if dm_m2c is not None:
            try:
                dm.M2C = dm_m2c  # Restore M2C into DM
            except Exception as e:
                logger.warning(f"Impossible to apply dm.M2C: {e}")  # Non-fatal

        results = {"arrays": {}, "scalars": {}, "figures": {}, "png_paths": []}  # Collected results
        results_dir = os.path.join(session_root, "results")  # Results subfolder (if not zipped)
        if os.path.isdir(results_dir):  # Optional: load if present (uncompressed case)
            for npy_path in sorted(glob.glob(os.path.join(results_dir, "*.npy"))):  # Arrays
                key = os.path.splitext(os.path.basename(npy_path))[0]
                try:
                    results["arrays"][key] = np.load(npy_path, allow_pickle=True)
                except Exception as e:
                    logger.warning(f"Impossible to load {npy_path}: {e}")
            scalars_path = os.path.join(results_dir, "scalars.json")  # Scalars
            if os.path.exists(scalars_path):
                try:
                    with open(scalars_path, "r") as f:
                        results["scalars"] = json.load(f)
                except Exception as e:
                    logger.warning(f"Impossible to load scalars.json: {e}")
            for pkl_path in sorted(glob.glob(os.path.join(results_dir, "*.fig.pkl"))):  # Figures
                key = os.path.basename(pkl_path).replace(".fig.pkl", "")
                try:
                    with open(pkl_path, "rb") as f:
                        import pickle  # Local import to avoid global dependency
                        results["figures"][key] = pickle.load(f)
                except Exception as e:
                    logger.warning(f"Impossible to load fig. {pkl_path}: {e}")
            results["png_paths"] = sorted(glob.glob(os.path.join(results_dir, "*.png")))  # PNG files list

        if cleanup_needed:  # Remove temp extraction dir if we created it
            try:
                shutil.rmtree(extracted_dir)
            except Exception:
                pass  # Best-effort cleanup

        return cl, results, session_root  # Return reconstructed objects and session root
    def load_data(self, zip_or_dir: str, keep_extracted: bool = False, load_figures: bool = True):
        """Load *only* metadata JSON and raw results from a saved session.

        Parameters
        ----------
        zip_or_dir : str
            Path to a session *.zip* produced by :meth:`save_all`, or an extracted session directory.
        keep_extracted : bool, optional
            If ``zip_or_dir`` is a zip file, keep the temporary extraction directory instead of deleting it.
        load_figures : bool, optional
            If True, unpickle ``*.fig.pkl`` Matplotlib figures found in the results folder.

        Returns
        -------
        tuple(dict, dict, str)
            (metadata_dict, results_dict, session_root_directory)

        Notes
        -----
        This method does **not** rebuild a :class:`CloseLoop` instance and does not load arrays such as
        IM/pupil/submask from ``metadata/``. It is intended for lightweight inspection or post-processing of
        scalar metadata and raw results.
        """
        # Detect whether we are dealing with a zip archive or a directory
        cleanup_needed = False
        if os.path.isdir(zip_or_dir):
            extracted_dir = zip_or_dir  # Use folder as-is
        else:
            tmpdir = tempfile.mkdtemp(prefix="oziriis_load_")  # Create temporary extraction directory
            with zipfile.ZipFile(zip_or_dir, "r") as zf:
                zf.extractall(tmpdir)  # Extract all session contents
            extracted_dir = tmpdir
            cleanup_needed = not keep_extracted

        # Locate the session root (typically the single timestamped subfolder)
        subdirs = [p for p in os.listdir(extracted_dir) if os.path.isdir(os.path.join(extracted_dir, p))]
        session_root = os.path.join(extracted_dir, subdirs[0]) if len(subdirs) == 1 else extracted_dir

        # Find metadata JSON (prefer the metadata/ subfolder, with backward compatibility)
        meta_dir = os.path.join(session_root, "metadata")
        meta_candidates = sorted(glob.glob(os.path.join(meta_dir, "*metadata*.json"))) if os.path.isdir(meta_dir) else []
        if not meta_candidates:
            meta_candidates = sorted(glob.glob(os.path.join(session_root, "**", "*_metadata_*.json"), recursive=True))
            if not meta_candidates:
                meta_candidates = sorted(glob.glob(os.path.join(session_root, "**", "*metadata*.json"), recursive=True))
            if not meta_candidates:
                if cleanup_needed:
                    try:
                        shutil.rmtree(extracted_dir)
                    except Exception:
                        pass
                raise FileNotFoundError("Metadata files not found in the session")
            meta_dir = session_root  # Old layout fallback
        meta_path = meta_candidates[-1]  # Pick the newest metadata file

        # Load metadata as a plain dict (unflattened)
        metadata = self._load_json(meta_path)

        # Collect results (arrays/scalars/figures/png paths) from results/ subfolder if present
        results = {"arrays": {}, "scalars": {}, "figures": {}, "png_paths": []}
        results_dir = os.path.join(session_root, "results")
        if os.path.isdir(results_dir):
            # Arrays (*.npy)
            for npy_path in sorted(glob.glob(os.path.join(results_dir, "*.npy"))):
                key = os.path.splitext(os.path.basename(npy_path))[0]
                try:
                    results["arrays"][key] = np.load(npy_path, allow_pickle=True)
                except Exception as e:
                    logger.warning(f"Failed to load array {npy_path}: {e}")
            # Scalars (scalars.json)
            scalars_path = os.path.join(results_dir, "scalars.json")
            if os.path.exists(scalars_path):
                try:
                    with open(scalars_path, "r") as f:
                        results["scalars"] = json.load(f)
                except Exception as e:
                    logger.warning(f"Failed to load scalars.json: {e}")
            # Figures (*.fig.pkl) – optional
            if load_figures:
                for pkl_path in sorted(glob.glob(os.path.join(results_dir, "*.fig.pkl"))):
                    key = os.path.basename(pkl_path).replace(".fig.pkl", "")
                    try:
                        import pickle
                        with open(pkl_path, "rb") as f:
                            results["figures"][key] = pickle.load(f)
                    except Exception as e:
                        logger.warning(f"Failed to load figure {pkl_path}: {e}")
            # PNGs (*.png)
            results["png_paths"] = sorted(glob.glob(os.path.join(results_dir, "*.png")))

        # Cleanup temporary extraction directory if applicable
        if cleanup_needed:
            try:
                shutil.rmtree(extracted_dir)
            except Exception:
                pass

        return metadata, results, session_root



